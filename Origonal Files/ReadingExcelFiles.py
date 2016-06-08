# -*- coding: utf-8 -*-
"""
Created on Fri Apr 29 18:04:13 2016

@author: japau
"""

import openpyxl
import pandas
from pprint import pprint
from datetime import datetime

dataFile  = r'2013_ERCOT_Hourly_Load_Data.xlsx'


def parse_file(datafile):
    workbook = openpyxl.load_workbook(datafile)
    worksheet = workbook.active
    numRows = worksheet.calculate_dimension()
    
    data = pandas.DataFrame([[cell.value for cell in col] for col in ws.iter_rows(numRows)])
    data.columns = data.iloc[0]
    data  = data.reindex(data.index.drop(0))    
    
    #print(len(data['COAST']))    
    
    cv = data['COAST']
    
    maxval = max(cv)
    #print (maxval)
    minval = min(cv)
    #print  (minval)
    
    maxpos = cv[cv == maxval].index[0] 
    #print (maxpos)
    minpos = cv[cv == minval].index[0]
    #print (minpos)
    
    maxtime = data['Hour_End'].iloc[maxpos]
    #print (maxtime)
    mintime = data['Hour_End'].iloc[maxpos]
    
    data = {'maxtime' : maxtime,
            'maxvalue' : maxval,
            'mintime' : mintime,
            'minvalue': minval,
            'avgcoast' : sum(cv) / float(len(cv))
            }
            
    return data




data = parse_file(dataFile)
pprint (data)
#workbook = openpyxl.load_workbook(dataFile)
#worksheet = workbook.active
#numRows = worksheet.calculate_dimension()
#
#data = pandas.DataFrame([[cell.value for cell in col] for col in ws.iter_rows(numRows)])
#data.columns = data.iloc[0]
#data  = data.reindex(data.index.drop(0))
#cv = data['COAST']  
#maxval = max(cv)
#print (maxval)
#minval = min(cv)
#print  (minval)
#maxpos = cv[cv == maxval].index[0]
#print (maxpos)
#maxtime = data['Hour_End'].iloc[maxpos]
#print (maxtime)          